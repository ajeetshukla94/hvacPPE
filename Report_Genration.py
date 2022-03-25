import pandas as pd
from flask import Flask, render_template, flash, url_for, request, make_response, jsonify, session, send_from_directory
from werkzeug.utils import secure_filename
import os, time
import io
import base64
import json
import datetime
import xlsxwriter
import os, sys, glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import Border, Side
# from flask import send_from_directory
from flask import send_file
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from fnmatch import fnmatch
import time

MYDIR = os.path.dirname(__file__)
#app.config['UPLOAD_FOLDER'] = "static/inputData/"

 
customer_details      = pd.read_excel(os.path.join("static/inputData/","company_details.xlsx"))
equipment_master      = pd.read_excel(os.path.join("static/inputData/","equipment_master.xlsx"))
EUGMP_guidlines       = pd.read_excel(os.path.join("static/inputData/","EUGMP_guidlines.xlsx"))
ISO_guidlines_master  = pd.read_excel(os.path.join("static/inputData/","ISO_guidlines.xlsx"))

equipment_master['SR_NO_ID'] = equipment_master['SR_NO_ID'].astype(str)
equipment_master['DUE_DATE'] = equipment_master['DUE_DATE'].astype(str)
equipment_master['DONE_DATE'] = equipment_master['DONE_DATE'].astype(str)

REGULAR_SIZE = 11
REGULAR_FONT = 'Cambria'
server = 'smtp.gmail.com'
port = 587
username = "aajeetshk@gmail.com"
password = "ilbumnmnsnqletdk"
send_from = "aajeetshk@gmail.com"
send_to = "ashish@pinpointengineers.co.in"


def sum_velocty(row):
    sum = int(row.V1) + int(row.V2) + int(row.V3) + int(row.V4) + int(row.V5)
    avg = sum / 5
    return int(avg)


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


class Report_Genration:
    def __init__(self):
        self = self

    @staticmethod
    def generate_report(data, basic_details,user):
        customer_details      = pd.read_excel(os.path.join("static/inputData/","company_details.xlsx"))
        sr_no = basic_details['sr_no']
        company_name = basic_details['company_name']
        room_volume = basic_details['room_volume']
        room_name = basic_details['room_name']
        ahu_number = basic_details['ahu_number']
        test_taken = basic_details['Test_taken']
        locatiom = basic_details['location']
        grade = basic_details['grade']
        acph_thresold = basic_details['acph_thresold']
        company_name_val = company_name
        SR_NO_val = sr_no
        test_date = test_taken

        temp_df = customer_details.loc[(customer_details.COMPANY_NAME == company_name_val)]
        report_number = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df = equipment_master.loc[(equipment_master.SR_NO_ID == SR_NO_val)]
        INSTRUMENT_NAME = temp_df.EQUIPMENT_NAME.values[0]
        MAKE = temp_df.MAKE.values[0]
        MAKE_MODEL = temp_df.MODEL_NUMBER.values[0]
        done_date = temp_df.DONE_DATE.values[0].split()[0]
        due_date = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test = "AIR VELOCITY"
        location = locatiom
        customer_name = company_name_val
        Test_taken = test_date
        ahu_number = ahu_number
        SRNO = SR_NO_val

        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")

        working_directory = MYDIR + "/" +"static/Report/AIR_VELOCITY_REPORT/{}"
        final_working_directory = "static/Report/AIR_VELOCITY_REPORT/{}/{}.xlsx"
        file_name = "{}_AIR_VELOCOTY_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));

        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR + "/" +final_working_directory.format(compan_name, file_name)
        print(final_working_directory)

        pattern = "*.xlsx"
        mypath     = 'static\\Report\\{}\{}'.format('AIR_VELOCITY_REPORT',compan_name)
        compare_date = datetime.datetime.today().strftime('%d/%m/%Y')
        print(compare_date)
        file_list= []
        for path, subdirs, files in os.walk(mypath):
            for name in files:
                if fnmatch(name, pattern):          
                    file_time = time.strftime('%d/%m/%Y',time.gmtime(os.path.getmtime(os.path.join(path, name) )))
                    datetime_object = datetime.datetime.strptime(file_time, '%d/%m/%Y').strftime('%d/%m/%Y')
                    print(datetime_object)
                    if  datetime_object==compare_date :
                        file_list.append(os.path.join(path, name))
        print("total file today")                
        print(len(file_list))
        
        report_number = "PPE0{}AV{}{}".format(report_number,datetime.datetime.today().strftime('%d%m%Y'),len(file_list)+1)
 
        wb = load_workbook(os.path.join("static/inputData/Template/",'Air_velocity_template.xlsx'))
        ws = wb.active
        ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        ws['F5'] = str(Nature_of_test)
        ws['F6'] = str(Test_taken)
        ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        ws['F9'] = str(done_date).replace("-", "/")
        ws['F10'] = str(due_date).replace("-", "/")
        ws['F11'] = report_number 

        ws['B17'] = str(INSTRUMENT_NAME)
        ws['E17'] = str(MAKE_MODEL)
        ws['I17'] = str(SRNO)
        ws['M17'] = str(VALIDITY.replace("-", "/"))

        ws.merge_cells(start_row=23, start_column=2, end_row=23, end_column=3)
        # ws.merge_cells(start_row=23, start_column=3, end_row=23, end_column=4)
        ws.merge_cells(start_row=23, start_column=15, end_row=23, end_column=16)

        row = 24
        data['AVG_Velocity'] = data.apply(sum_velocty, axis=1)
        data['CFM'] = data.apply(lambda x: int(x['AVG_Velocity']) * int(x['Inlet_size']), axis=1)
        Total_cfm = data['CFM'].sum()
        ACPH_VALUE = float((Total_cfm * 60)) / float(room_volume)
        ACPH_VALUE = int(ACPH_VALUE)

        # set_border(ws, "B16:M17")
        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            ws["D" + str(row)] = row_data.Label_number
            currentCell = ws["D" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')
            ws["F" + str(row)] = row_data.V1
            ws["G" + str(row)] = row_data.V2
            ws["H" + str(row)] = row_data.V3
            ws["I" + str(row)] = row_data.V4
            ws["J" + str(row)] = row_data.V5
            ws["K" + str(row)] = row_data.AVG_Velocity
            ws["L" + str(row)] = row_data.Inlet_size
            ws["M" + str(row)] = row_data.CFM
            row += 1

        ws.merge_cells(start_row=24, start_column=2, end_row=row - 1, end_column=3)
        ws['B24'] = room_name

        ws.merge_cells(start_row=24, start_column=14, end_row=row - 1, end_column=14)
        ws['N24'] = room_volume

        ws.merge_cells(start_row=24, start_column=15, end_row=row - 1, end_column=16)
        ws['O24'] = ACPH_VALUE

        currentCell = ws['B24']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        currentCell = ws['N24']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        currentCell = ws['O24']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        ws["B" + str(row)] = "TOTAL CFM"

        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=13)
        ws["M" + str(row)] = Total_cfm
        ws.merge_cells(start_row=row, start_column=15, end_row=row, end_column=16)

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        ws["B" + str(row)] = "Acceptance criteria :Not less Than {} ACPH".format(str(acph_thresold))

        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=16)
        ws["J" + str(row)] = "GRADE : {}".format(grade)

        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 2)] = "TEST CARRIED OUT BY"

        currentCell = ws["B" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 2)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEER"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)

        ws["B" + str(row + 9)] = "Test Carried out by {} on {}".format(user,datetime.datetime.today().strftime('%d/%m/%Y %H:%M:%S'))



        set_border(ws, 'B1:P' + str(row + 9))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)

        return file_name, store_location

    @staticmethod
    def generate_report_pao(data, basic_details,user):
        customer_details      = pd.read_excel(os.path.join("static/inputData/","company_details.xlsx"))
        sr_no = basic_details['sr_no']
        company_name = basic_details['company_name']
        room_name = basic_details['room_name']
        ahu_number = basic_details['ahu_number']

        test_taken = basic_details['Test_taken']
        locatiom = basic_details['location']
        compresed_value = basic_details['compresed_value']
        check_val = basic_details['check_val']
        company_name_val = company_name
        SR_NO_val = sr_no
        test_date = test_taken

        temp_df = customer_details.loc[(customer_details.COMPANY_NAME == company_name_val)]
        report_number = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df = equipment_master.loc[(equipment_master.SR_NO_ID == SR_NO_val)]
        INSTRUMENT_NAME = temp_df.EQUIPMENT_NAME.values[0]
        MAKE = temp_df.MAKE.values[0]
        MAKE_MODEL = temp_df.MODEL_NUMBER.values[0]
        done_date = temp_df.DONE_DATE.values[0].split()[0]
        due_date = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test = "PAO REPORT"
        location = locatiom
        customer_name = company_name_val
        Test_taken = test_date
        SRNO = SR_NO_val

        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")

        working_directory = MYDIR + "/" "static/Report/PAO_REPORT/{}"
        final_working_directory = "static/Report/PAO_REPORT/{}/{}.xlsx"
        file_name = "{}_PAO_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));

        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR + "/"+final_working_directory.format(compan_name, file_name)
        print(final_working_directory)
        
        
        pattern = "*.xlsx"
        mypath     = 'static\\Report\\{}\{}'.format('PAO_REPORT',compan_name)
        compare_date = datetime.datetime.today().strftime('%d/%m/%Y')
        print(compare_date)
        file_list= []
        for path, subdirs, files in os.walk(mypath):
            for name in files:
                if fnmatch(name, pattern):          
                    file_time = time.strftime('%d/%m/%Y',time.gmtime(os.path.getmtime(os.path.join(path, name) )))
                    datetime_object = datetime.datetime.strptime(file_time, '%d/%m/%Y').strftime('%d/%m/%Y')
                    print(datetime_object)
                    if  datetime_object==compare_date :
                        file_list.append(os.path.join(path, name))
        
        report_number = "PPE0{}PA{}{}".format(report_number,datetime.datetime.today().strftime('%d%m%Y'),len(file_list)+1)

        wb = load_workbook(os.path.join("static/inputData/Template/",'PAO_template.xlsx'))
        ws = wb.active
        ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        ws['F5'] = str(Nature_of_test)
        ws['F6'] = str(Test_taken)
        ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        ws['F9'] = str(done_date).replace("-", "/")
        ws['F10'] = str(due_date).replace("-", "/")
        ws['F11'] = report_number#str("PPE0{}AV01A".format(report_number))

        ws['B17'] = str(INSTRUMENT_NAME)
        ws['E17'] = str(MAKE_MODEL)
        ws['I17'] = str(SRNO)
        ws['M17'] = str(VALIDITY.replace("-", "/"))

        row = 24

        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            ws["D" + str(row)] = row_data.INLET_NUMBER
            currentCell = ws["D" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
            ws["F" + str(row)] = row_data.Upstream
            currentCell = ws["F" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
            ws["I" + str(row)] = row_data.Leakage
            currentCell = ws["I" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
            ws["L" + str(row)] = row_data.Remark
            currentCell = ws["L" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        ws.merge_cells(start_row=24, start_column=2, end_row=row - 1, end_column=3)
        ws['B24'] = room_name

        currentCell = ws['B24']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        ws["B" + str(row)] = "Compressed Air Presure {}".format(compresed_value)
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        ws["B" + str(row)] = "Acceptance criteria :Not less Than {} %".format(str(check_val))

        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 2)] = "TEST CARRIED OUT BY"

        currentCell = ws["B" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 2)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEER"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)
        ws["B" + str(row + 9)] = "Test Carried out by {} on {}".format(user,datetime.datetime.today().strftime('%d/%m/%Y %H:%M:%S'))


        set_border(ws, 'B1:P' + str(row + 9))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)

        return file_name, store_location

    @staticmethod
    def generate_report_particle_count(data, basic_details,user):
        customer_details      = pd.read_excel(os.path.join("static/inputData/","company_details.xlsx"))
        sr_no = basic_details['sr_no']
        company_name = basic_details['company_name']
        room_name = basic_details['room_name']
        ahu_number = basic_details['ahu_number']
        test_taken = basic_details['Test_taken']
        locatiom = basic_details['location']
        condition = basic_details['condition']
        grade = basic_details['grade']
        gl_value = basic_details['gl_value']
        company_name_val = company_name
        SR_NO_val = sr_no
        test_date = test_taken

        temp_df = customer_details.loc[(customer_details.COMPANY_NAME == company_name_val)]
        report_number = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df = equipment_master.loc[(equipment_master.SR_NO_ID == SR_NO_val)]
        INSTRUMENT_NAME = temp_df.EQUIPMENT_NAME.values[0]
        MAKE = temp_df.MAKE.values[0]
        MAKE_MODEL = temp_df.MODEL_NUMBER.values[0]
        done_date = temp_df.DONE_DATE.values[0].split()[0]
        due_date = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test = "PARTICLE COUNT REPORT"
        location = locatiom
        customer_name = company_name_val
        Test_taken = test_date
        SRNO = SR_NO_val


        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")
        print(gl_value)
        
        if "ISO" in gl_value :    
            trn    = ISO_guidlines_master.loc[  (ISO_guidlines_master['Grade']== grade)  ]    
        if "EU" in gl_value  :
            trn = EUGMP_guidlines[(EUGMP_guidlines['Condition'] == condition) & (EUGMP_guidlines['Grade'] == grade)]
        print(trn)
        working_directory = MYDIR + "/"  "static/Report/PARTICLE_REPORT/{}"
        final_working_directory = "static/Report/PARTICLE_REPORT/{}/{}.xlsx"
        file_name = "{}_PARTICLE_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));
        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR +"/" + final_working_directory.format(compan_name, file_name)
        print(final_working_directory)
        
        pattern = "*.xlsx"
        mypath     = 'static\\Report\\{}\{}'.format('PARTICLE_REPORT',compan_name)
        compare_date = datetime.datetime.today().strftime('%d/%m/%Y')
        print(compare_date)
        file_list= []
        for path, subdirs, files in os.walk(mypath):
            for name in files:
                if fnmatch(name, pattern):          
                    file_time = time.strftime('%d/%m/%Y',time.gmtime(os.path.getmtime(os.path.join(path, name) )))
                    datetime_object = datetime.datetime.strptime(file_time, '%d/%m/%Y').strftime('%d/%m/%Y')
                    print(datetime_object)
                    if  datetime_object==compare_date :
                        file_list.append(os.path.join(path, name))
        
        report_number = "PPE0{}PC{}{}".format(report_number,datetime.datetime.today().strftime('%d%m%Y'),len(file_list)+1)

        wb = load_workbook(os.path.join("static/inputData/Template/",'particle_count_template.xlsx'))
        ws = wb.active
        ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        ws['F5'] = str(Nature_of_test)
        ws['F6'] = str(Test_taken)
        ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        ws['F9'] = str(done_date).replace("-", "/")
        ws['F10'] = str(due_date).replace("-", "/")
        ws['F11'] = report_number#str("PPE0{}AV01A".format(report_number))

        ws['B17'] = str(INSTRUMENT_NAME)
        ws['E17'] = str(MAKE_MODEL)
        ws['I17'] = str(SRNO)
        ws['M17'] = str(VALIDITY.replace("-", "/"))

        row = 23

        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
            ws["F" + str(row)] = row_data.Location
            currentCell = ws["F" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
            ws["H" + str(row)] = row_data.zeor_point_five
            currentCell = ws["H" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
            ws["J" + str(row)] = row_data.five_point_zero
            currentCell = ws["J" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
            ws["L" + str(row)] = row_data.remark
            currentCell = ws["L" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        ws.merge_cells(start_row=23, start_column=4, end_row=row - 1, end_column=5)
        ws['D23'] = room_name
        ws.merge_cells(start_row=23, start_column=2, end_row=row - 1, end_column=3)
        ws['B23'] = "1"
        currentCell = ws['B23']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        currentCell = ws['D23']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        data.zeor_point_five = data.zeor_point_five.astype("float")
        data.five_point_zero = data.five_point_zero.astype("float")

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws["F" + str(row)] = "Average"
        currentCell = ws["F" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws["H" + str(row)] = data.zeor_point_five.mean()
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        ws["J" + str(row)] = data.five_point_zero.mean()
        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws["F" + str(row)] = "STD"
        currentCell = ws["F" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws["H" + str(row)] = data.zeor_point_five.std()
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        ws["J" + str(row)] = data.five_point_zero.std()
        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=16)
        ws["B" + str(row)] = "Maximum {}".format(gl_value)
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        row = row + 2

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws["B" + str(row)] = "GRADE"
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
        ws["H" + str(row)] = " 0.5 m"
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        ws["L" + str(row)] = " 5.0 m"
        currentCell = ws["L" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws["B" + str(row)] = "GRADE {}".format(grade.upper())
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
        ws["H" + str(row)] = trn['point_five_percent'].values[0]
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        ws["L" + str(row)] = trn['five_percent'].values[0]
        currentCell = ws["L" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 1)] = "TEST CARRIED OUT BY"

        currentCell = ws["B" + str(row + 1)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 1)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 1)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEER"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)

        ws["B" + str(row + 9)] = "Test Carried out by {} on {}".format(user,datetime.datetime.today().strftime('%d/%m/%Y %H:%M:%S'))

        set_border(ws, 'B1:P' + str(row + 8))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)

        return file_name, store_location




